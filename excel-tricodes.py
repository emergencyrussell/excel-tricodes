import random
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Step 1: Read the two text files into two lists
with open('trigrams.txt', 'r') as file:
    list1 = file.read().splitlines()

with open('codedwords.txt', 'r') as file:
    list2 = file.read().splitlines()

# Ask the user for a seed
seed = input("Please enter seed: ")
if seed:
    random.seed(seed)
else:
    random.seed(0)

# Set the number of encode/decode pairs based on command-line argument
if len(sys.argv) > 1:
    try:
        num_pairs = int(sys.argv[1])
    except ValueError:
        print("Invalid input. Please enter a number.")
        sys.exit(1)
else:
    num_pairs = 1  # Default number of pairs

# Function to write pairs to an Excel sheet
def write_pairs_to_excel(ws, pairs, sort_key, font):
    pairs = sorted(pairs, key=sort_key)
    chunk_size = pairs_per_chunk * horizontal_chunks
    row_offset = 0
    for start_idx in range(0, len(pairs), chunk_size):  # Process by chunk_size
        chunk_pairs = pairs[start_idx:start_idx + chunk_size]

        for j in range(horizontal_chunks):  # Number of horizontal chunks
            column_pairs = chunk_pairs[j * pairs_per_chunk:(j + 1) * pairs_per_chunk]

            for i, pair in enumerate(column_pairs):
                col_base = j * 2 + 1
                cell1 = ws.cell(row=row_offset + i + 1, column=col_base, value=pair[0])
                cell2 = ws.cell(row=row_offset + i + 1, column=col_base + 1, value=pair[1])
                cell1.font = font
                cell2.font = font

        row_offset += pairs_per_chunk + 1  # Move to the next block of pairs_per_chunk rows (plus one blank row)

    # Autofit columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

# Initialize workbook
wb = Workbook()
font = Font(name='Courier New', b=True, sz=10.1)

# Set the number of pairs in a chunk and the number of horizontal chunks
pairs_per_chunk = 5
horizontal_chunks = 3

# Create specified number of sheet pairs
for pair_num in range(num_pairs):
    # Randomly pair elements from list1 and list2 for each pair of sheets
    indices1 = random.sample(range(len(list1)), len(list1))
    indices2 = random.sample(range(len(list2)), len(list2))

    pairs = list(zip(
        (list1[i] for i in indices1),
        (list2[i] for i in indices2)
    ))

    # Truncate the pairs to match the length of the shorter list
    min_length = min(len(list1), len(list2))
    pairs = pairs[:min_length]

    # Create Encode sheet
    encode_sheet = wb.create_sheet(title=f'Encode {pair_num + 1}')
    write_pairs_to_excel(encode_sheet, pairs, lambda x: x[0].lower(), font)

    # Create Decode sheet
    decode_sheet = wb.create_sheet(title=f'Decode {pair_num + 1}')
    write_pairs_to_excel(decode_sheet, pairs, lambda x: x[1].lower(), font)

# Remove the default sheet created by Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the workbook
wb.save('tricodes.xlsx')
